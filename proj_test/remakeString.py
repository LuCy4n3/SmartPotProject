import openpyxl
import os

pathExcl = "c:/Users/simon/OneDrive/Desktop/proj_test/linksForPlants.xlsx"
workbook = openpyxl.load_workbook(pathExcl)
sheet = workbook.get_sheet_by_name('completed')
if(not sheet):
    print('error in excel')
    exit(-1)
else :
    print("got data")

for row in  sheet.iter_rows(0,sheet.max_row):
    new_value = row[1].value[:4] + 's' + row[1].value[4:]
    print(new_value)
    row[1].value = new_value
workbook.save(pathExcl)
workbook.close()
print("Data updated successfully")  
    
'''
for index, row in sheet.iter_rows(0,sheet.max_row-2):
    print(row)
    new_value = row[1].value[:4] + 's' + row[1].value[4:]
    print(new_value)
    #sheet.cell(row=index, column=2).value = x1   
'''