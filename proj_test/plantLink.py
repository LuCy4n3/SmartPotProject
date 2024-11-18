import asyncio
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
import time
import random
import openpyxl

listOfNumb = [2, 3, 4, 5, 6, 7]

workbook = openpyxl.Workbook()
# Select the default sheet (usually named 'Sheet')
sheetCompl = workbook.create_sheet("completed",1)
sheetFailed = workbook.create_sheet("failed",2)
sheetPlants = workbook.create_sheet("plants",3)



async def new_page(link, browser, retry_attempts=3):
    context = await browser.new_context()

    await context.add_cookies([
        {
            'name': 'cusess',
            'value': '98291da5e63b0067f2bf7f9b80b69529',
            'domain': 'garden.org',
            'path': '/'
        },
        {
            'name': 'sr',
            'value': '%2F',
            'domain': 'garden.org',
            'path': '/'
        },
        {
            'name': '_pk_id.1.94cf',
            'value': '3f0c64ea55e7161b.1729536175.',
            'domain': 'garden.org',
            'path': '/'
        },
        {
            'name': 'cf_clearance',
            'value': 'Xyxrc0sYB8J40hqfhpcFw2jk6wYv3qk8rBZZLOxBTn0-1729538939-1.2.1.1-jTU_oct3Tt.75ZDClaOgm5v1_IIje8MZcLzm9eey.c1UTsNKvtrseaNMy5bQaQdNXkWvwflP7je_zQqJCe3DvjU2LEEQC9s6JxCspIkGhauc3D41GVJo3kFqGCOsgJfpSu57MpQUpaRdhHVbFU3IGIdseD0xEGdC.jGcU8CZyXxseKQn2dnH8Imf7.OiYWSPj9j3LR3697NLK0hyectEy8LLbdBMrwKPXM7k..0p6zz6kpnv3EuCZTUKwich9pIxeRXNpB4OMgCh9.6GStZZJAhVgTmkotps26qo9wqvKzwaCKa81D848PQZLuZI0Nc.yWQwzB6J_qTfHRD9z1UFG6.jB_FZ4LgCNawxcLFzWizdPp_stgObhrXGTYbhaRAO',
            'domain': 'garden.org',
            'path': '/'
        }
    ])

    page = await context.new_page()

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'en,en-US;q=0.9,ro;q=0.8,de;q=0.7',
        'Connection': 'keep-alive',
        'Referer': 'https://medium.com/zenrows/web-scraping-without-getting-blocked-cbafa55d8045',
        'Sec-Ch-Ua': '"Google Chrome";v="129", "Not=A?Brand";v="8", "Chromium";v="129"',
        'Sec-Ch-Ua-Mobile': '?0',
        'Sec-Ch-Ua-Platform': '"Windows"',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'cross-site',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1'
    }

    await page.set_extra_http_headers(headers)

    attempt = 0
    while attempt < retry_attempts:
        try:
            await asyncio.wait_for(page.goto("https://garden.org" + link), timeout=5)
            content = await page.content()
            await page.close()
            await context.close()
            return content
        except asyncio.TimeoutError:
            attempt += 1
            retry_delay = random.uniform(1, 3)
            print(f"Timeout reached. Retrying in {retry_delay:.2f} seconds...")
            await asyncio.sleep(retry_delay)
        except Exception as e:
            print(f"Exception occurred: {e}. Retrying...")
            attempt += 1
            await page.close()
            page = await context.new_page()  # reopen page for new attempt

    await page.close()
    await context.close()
    print(f"Failed to fetch {link} after {retry_attempts} attempts.")
    return None

async def fetch_links(content, link):
    if content is None:
        print(f"Skipping link due to failed fetch: {link['href']}")
        return

    print("Processing content for link:", "https://garden.org" + link['href'])
    soup = BeautifulSoup(content, "html.parser")
    holders = soup.find("div", {"id": "_search_and_browse"})
    if holders:
        divLink = holders.find_next('div').find("div", {"class": "card-body"})
        if divLink:
            inner_link = divLink.find('a')
            if inner_link:
                print(inner_link['href'])
                currentSheet = workbook.get_sheet_by_name('completed')
                currentSheet.append([link.get_text(),"https://garden.org" + inner_link['href']])
                #compl.write("https://garden.org" +inner_link['href']+"\n")
    else:
        print(f"No '_search_and_browse' section found for link: {link['href']}")

async def main():
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()

        await context.add_cookies([
            {
                'name': 'cusess',
                'value': '98291da5e63b0067f2bf7f9b80b69529',
                'domain': 'garden.org',
                'path': '/'
            },
            {
                'name': 'sr',
                'value': '%2F',
                'domain': 'garden.org',
                'path': '/'
            },
            {
                'name': '_pk_id.1.94cf',
                'value': '3f0c64ea55e7161b.1729536175.',
                'domain': 'garden.org',
                'path': '/'
            },
            {
                'name': 'cf_clearance',
                'value': 'Xyxrc0sYB8J40hqfhpcFw2jk6wYv3qk8rBZZLOxBTn0-1729538939-1.2.1.1-jTU_oct3Tt.75ZDClaOgm5v1_IIje8MZcLzm9eey.c1UTsNKvtrseaNMy5bQaQdNXkWvwflP7je_zQqJCe3DvjU2LEEQC9s6JxCspIkGhauc3D41GVJo3kFqGCOsgJfpSu57MpQUpaRdhHVbFU3IGIdseD0xEGdC.jGcU8CZyXxseKQn2dnH8Imf7.OiYWSPj9j3LR3697NLK0hyectEy8LLbdBMrwKPXM7k..0p6zz6kpnv3EuCZTUKwich9pIxeRXNpB4OMgCh9.6GStZZJAhVgTmkotps26qo9wqvKzwaCKa81D848PQZLuZI0Nc.yWQwzB6J_qTfHRD9z1UFG6.jB_FZ4LgCNawxcLFzWizdPp_stgObhrXGTYbhaRAO',
                'domain': 'garden.org',
                'path': '/'
            }
        ])

        page = await context.new_page()
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'en,en-US;q=0.9,ro;q=0.8,de;q=0.7',
            'Connection': 'keep-alive',
            'Referer': 'https://medium.com/zenrows/web-scraping-without-getting-blocked-cbafa55d8045',
            'Sec-Ch-Ua': '"Google Chrome";v="129", "Not=A?Brand";v="8", "Chromium";v="129"',
            'Sec-Ch-Ua-Mobile': '?0',
            'Sec-Ch-Ua-Platform': '"Windows"',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'cross-site',
            'Sec-Fetch-User': '?1',
            'Upgrade-Insecure-Requests': '1'
        }

        await page.set_extra_http_headers(headers)
        await page.goto("https://garden.org/plants/browse/plants/children/181473/")

        content = await page.content()
        await page.close()
        soup = BeautifulSoup(content, "html.parser")
        print('got data')
        print(content)
        time.sleep(0.9)
        await context.close()
        
        if soup.get_text():
            for caption in soup.find_all('table'):
                for links in caption.find_all('a'):
                    print(links['href'])
                    time.sleep(random.choice(listOfNumb) / 10)
                    content = await new_page(links['href'], browser)
                    if content :
                        time.sleep(random.choice(listOfNumb)/5)
                        await fetch_links(content,links)
                        print(links.get_text() + " http://garden.org" + links['href'])
                        currentSheet = workbook.get_sheet_by_name('plants')
                        currentSheet.append([links.get_text(),"http://garden.org"+links['href']])
                        #plants.write(links.get_text()+"\n")
                    else :
                        print('got timeout')
                        currentSheet = workbook.get_sheet_by_name('failed')
                        currentSheet.append([links.get_text(),"http://garden.org" + links['href']])
                        #failed.write("http://garden.org" + links['href']+"\n")
                        exit
        await browser.close()
        #compl.close()
        #failed.close()
        #plants.close()
        workbook.save("linksForPlants.xlsx")
        print("file saved!")


asyncio.run(main())
