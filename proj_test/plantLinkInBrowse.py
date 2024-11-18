import asyncio
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
import time
import random
import openpyxl
import os
import google.generativeai as genai
import json

pathExcl = "c:/Users/simon/OneDrive/Desktop/proj_test/linksForPlants.xlsx"

# used / instead of \ bcs im trying to use a normal string as a path
workbook = openpyxl.load_workbook(pathExcl)
allLinksWorkbook = openpyxl.Workbook()

listOfNumb = [2, 3, 4, 5, 6, 7]


plants = allLinksWorkbook.create_chartsheet("plantsAndSpecies",1)


#gemini AI chat
# can be use with the command: chat_session.send_message("enter prompt")
#prompt for the following history "for the same requests as said before tell me the params for <urPlant>"
genai.configure(api_key="AIzaSyDSNuqPbWMoHLbcyyliNbJsYLgxNtKKau4")

# Create the model
generation_config = {
  "temperature": 2,
  "top_p": 0.95,
  "top_k": 40,
  "max_output_tokens": 8192,
  "response_mime_type": "text/plain",
}

model = genai.GenerativeModel(
  model_name="gemini-1.5-flash-002",
  generation_config=generation_config,
)

chat_session = model.start_chat(
  history=[
    {
      "role": "user",
      "parts": [
        "what are the growth parameters for Daylily (Hemerocallis citrina). I want u to give me the answer as simple as u can in the following fields: water preference, life cycle, plant habit, flower color, pH min, pH max, minim temperature, max temperature, sun req, plant height, plant width, fruiting time, flower time, soil type, NPK values, spacing, humidity and the plant name which is just the species of plant. Give the response just as a JSON object with all the params. The names will be:PlantName,PlantGroup,WaterPref,LifeCycle,PlantHabit,FlowerColor,PhMinVal,PhMaxVal,MinTemp,MaxTemp,SunReq,PlantHeight,PlantWidth,FruitingTime,FlowerTime,SoilType,Nitrogen,Phosphorus,Potassium,Spacing,Humidity. Give me the answers for the distances, temps, and humidity in metric without adding the measurement. For distances leave only the max value. For NPK add the values in percentages for solutions per chemical agent and don't add \"%\". Set the humidity in percentage value choose the avg. For the sun req make it in hours. Dont add comments or suggestions, keep only the JSON body. Keep all the JSON params that use strings as simple as you can.",
      ],
    },
    {
      "role": "model",
      "parts": [
        "```json\n{\"PlantName\": \"Hemerocallis citrina\", \"PlantGroup\": \"Daylily\", \"WaterPref\": \"Moderate\", \"LifeCycle\": \"Perennial\", \"PlantHabit\": \"Herbaceous\", \"FlowerColor\": \"Yellow\", \"PhMinVal\": \"6\", \"PhMaxVal\": \"7\", \"MinTemp\": \"-20\", \"MaxTemp\": \"35\", \"SunReq\": \"6\", \"PlantHeight\": \"60\", \"PlantWidth\": \"30\", \"FruitingTime\": \"Summer\", \"FlowerTime\": \"Summer\", \"SoilType\": \"Well-drained loam\", \"Nitrogen\": \"10\", \"Phosphorus\": \"10\", \"Potassium\": \"10\", \"Spacing\": \"30\", \"Humidity\": \"60\"}\n\n```",
      ],
    },
  ]
)



def parse_and_serialize_json(json_text):
    try:
        

        json_text = json_text.strip()
        json_text = json_text.lstrip("```json")
        json_text = json_text.rstrip("```")  

        #print(json_text)
        data = json.loads(json_text)
        
        serialized_data = json.dumps(data, indent=2)
        
        return serialized_data
    
    except json.JSONDecodeError:
        print("=========================")
        print(json_text)
        print("===============================")
        return "Invalid JSON format"


async def fetch_nmbOfPages(content, link):
    if content is None:
        print(f"Skipping link due to failed fetch: {link}")
        return

    print("Getting the nmb of pages...")
    soup = BeautifulSoup(content, "html.parser")
    navBar = soup.find("nav", {"aria-label": "Pages"})
    print(navBar)
    if navBar:
        pageNmbrs = navBar.find_all('li')
        if pageNmbrs:
            for page in pageNmbrs:
                pass
            return page.get_text()
    else:
        print(f"No section found for link: {link}")
        return -1

async def new_page(link, browser, retry_attempts=3):
    context = await browser.new_context()

    await context.add_cookies([
        {
            'name': 'cusess',
            'value': '98291da5e63b0067f2bf7f9b80b69529',
            'domain': 'garden.org',
            'path': '/'
        },
        {
            'name': 'sr',
            'value': '%2F',
            'domain': 'garden.org',
            'path': '/'
        },
        {
            'name': '_pk_id.1.94cf',
            'value': '3f0c64ea55e7161b.1729536175.',
            'domain': 'garden.org',
            'path': '/'
        },
        {
            'name': 'cf_clearance',
            'value': 'Xyxrc0sYB8J40hqfhpcFw2jk6wYv3qk8rBZZLOxBTn0-1729538939-1.2.1.1-jTU_oct3Tt.75ZDClaOgm5v1_IIje8MZcLzm9eey.c1UTsNKvtrseaNMy5bQaQdNXkWvwflP7je_zQqJCe3DvjU2LEEQC9s6JxCspIkGhauc3D41GVJo3kFqGCOsgJfpSu57MpQUpaRdhHVbFU3IGIdseD0xEGdC.jGcU8CZyXxseKQn2dnH8Imf7.OiYWSPj9j3LR3697NLK0hyectEy8LLbdBMrwKPXM7k..0p6zz6kpnv3EuCZTUKwich9pIxeRXNpB4OMgCh9.6GStZZJAhVgTmkotps26qo9wqvKzwaCKa81D848PQZLuZI0Nc.yWQwzB6J_qTfHRD9z1UFG6.jB_FZ4LgCNawxcLFzWizdPp_stgObhrXGTYbhaRAO',
            'domain': 'garden.org',
            'path': '/'
        }
    ])

    page = await context.new_page()

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'en,en-US;q=0.9,ro;q=0.8,de;q=0.7',
        'Connection': 'keep-alive',
        'Referer': 'https://medium.com/zenrows/web-scraping-without-getting-blocked-cbafa55d8045',
        'Sec-Ch-Ua': '"Google Chrome";v="129", "Not=A?Brand";v="8", "Chromium";v="129"',
        'Sec-Ch-Ua-Mobile': '?0',
        'Sec-Ch-Ua-Platform': '"Windows"',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'cross-site',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1'
    }

    await page.set_extra_http_headers(headers)

    attempt = 0
    while attempt < retry_attempts:
        try:
            #the link is already formed as http:/..../header
            print(f"accesing {link} page")
            await asyncio.wait_for(page.goto(link), timeout=5)
            content = await page.content()
            await page.close()
            await context.close()
            return content
        except asyncio.TimeoutError:
            attempt += 1
            retry_delay = random.uniform(1, 3)
            print(f"Timeout reached. Retrying in {retry_delay:.2f} seconds...")
            await asyncio.sleep(retry_delay)
        except Exception as e:
            print(f"Exception occurred: {e}. Retrying...")
            attempt += 1
            await page.close()
            page = await context.new_page()  # reopen page for new attempt

    await page.close()
    await context.close()
    print(f"Failed to fetch {link} after {retry_attempts} attempts.")
    return None

async def fetch_data(content,plantName):
    if content is None:
        print("The contet provided is null")
        return 
    print(f"looking for data for {plantName}...")
    

async def fetch_links(content, link):
    if content is None:
        print(f"Skipping link due to failed fetch: {link}")
        return
    print("Processing content for link:", link)
    soup = BeautifulSoup(content, "html.parser")
    holders = soup.find("table", {"class": "table table-striped table-bordered table-hover caption-top pretty-table"})
    if holders:
        inner_link = holders.findAll('tr')
        finalPlant = ''
        for plant in inner_link:
            if plant:
                #print(plant)
                tableRowLink = plant.find('a')

                print(plant.get_text())

                finalPlant = finalPlant +', '+ plant.get_text()
                #respone = chat_session.send_message("for the same requests as said before tell me the params for "+plant.get_text())
                #print(parse_and_serialize_json(respone.text))
                #currentSheet = workbook.get_sheet_by_name('completed')
                #compl.write("https://garden.org" +inner_link['href']+"\n")
        
        respone = chat_session.send_message("for the same requests as said before tell me the params for "+finalPlant+". I want for each of those plants a JSON with the parms and all those JSON put in a big JSON array")
        print(parse_and_serialize_json(respone.text))
    else:
        print(f"No section found for link: {link}")
        print(soup)


async def main():
    async with async_playwright() as p:

        sheet = workbook.get_sheet_by_name('completed')
        if(not sheet):
            print('error in excel')
            exit(-1)
        else :
            print("got data")

        browser = await p.chromium.launch(headless=True)
        #print(sheet.max_column)
        for row in  sheet.iter_rows(0,sheet.max_row):
            #print(row[0].value+"  "+row[1].value)
            plantGroup = row[0].value
            plantGroupLink = row[1].value
            time.sleep(random.choice(listOfNumb)/5)
            content = await new_page(plantGroupLink,browser)
            if content :
                time.sleep(random.choice(listOfNumb)/5)
                nmbOfPages = await fetch_nmbOfPages(content,plantGroupLink)
                if int(nmbOfPages) -1 >= 0:
                    time.sleep(random.choice(listOfNumb)/5)
                    for i in range(1,int(nmbOfPages)+1):
                        pagePos = "?offset=" + str((i-1)*20)
                        print(plantGroupLink+pagePos)
                        await fetch_links(content,plantGroupLink+pagePos)
                        new_folder_path = "c:/Users/simon/OneDrive/Desktop/proj_test/plants/"+plantGroup
                        try:
                            # Create the folder
                            os.mkdir(new_folder_path)
                            print("Folder created:", new_folder_path)
                        except FileExistsError:
                            print("Folder already exists. Recreating:",new_folder_path)
                            os.rmdir(new_folder_path)
                            os.mkdir(new_folder_path)
                        except FileNotFoundError:
                            print("The parent directory does not exist.")
                            exit(-1)
                        except PermissionError:
                            print("Permission denied to create the folder.")
                            exit(-1)
                else :
                    print("not enough pages!!")

        



        #new_page('test',browser)
        
asyncio.run(main())


        
